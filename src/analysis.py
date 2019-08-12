import openpyxl
from openpyxl.utils.cell import get_column_letter

wb = openpyxl.load_workbook("excel_output/PM25_2019_05.xlsx")
rawdata = wb['Original Data']
avg_3h = wb['3h Average']
regionmax = wb['Regional Hour Max']
everything = wb['Everything Together']
# og data
for r in range(1, rawdata.max_row + 1):
    for c in range(1, rawdata.max_column + 1):
        everything[get_column_letter(c) + str(r)] = '=IF((\'Original Data\'!' + get_column_letter(c) + str(
            r) + ')="","",\'Original Data\'!' + get_column_letter(c) + str(r) + ')'

# '=(\'3h Average\'!' + get_column_letter(avg_C) + str(avg_R) + ')'
delta = rawdata.max_column + 1
for avg_C in range(3, avg_3h.max_column + 1):
    for avg_R in range(1, avg_3h.max_row + 1):
        everything[get_column_letter(avg_C + delta) + str(avg_R)] = '=IF((\'3h Average\'!' + get_column_letter(
            avg_C) + str(
            avg_R) + ')="","",\'3h Average\'!' + get_column_letter(avg_C) + str(avg_R) + ')'

# '=(\'Regional Hour Max\'!' + get_column_letter(reg_C) + str(reg_R) + ')'
avg3h_delta = avg_3h.max_column
for reg_C in range(3, regionmax.max_column + 1):
    for reg_R in range(1, regionmax.max_row + 1):
        everything[get_column_letter(reg_C + delta + avg3h_delta) + str(
            reg_R)] = '=IF((\'Regional Hour Max\'!' + get_column_letter(reg_C) + str(
            reg_R) + ')="","",\'Regional Hour Max\'!' + get_column_letter(reg_C) + str(reg_R) + ')'

wb.save("excel_output/PM25_2019_05.xlsx")
