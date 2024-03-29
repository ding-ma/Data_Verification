import csv
import datetime
import os
import shutil
from collections import OrderedDict

import openpyxl
import pandas as pd
from openpyxl.chart import (
    ScatterChart,
    Series,
    Reference,
)
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill, NamedStyle
from openpyxl.utils.cell import get_column_letter

######################################################
# to change conditional formatting, change the upper and lower bounds

# PM2.5
# above 50ppb
greaterOrEqual_PM25 = ['50.00001']

# between 35 and 50 ppb
secondHighest_PM25 = ['35.00001', '50']

# between 30 and 35 ppb
thirdHighest_PM25 = ['30', '35']

# between 25 and 30 ppb
lowest_PM25 = ['25.00001', '29.9999']
# end of PM2.5


# O3
greaterOrEqual_O3 = ['100.']
secondHighest_O3 = ['82', '99.9999']
thirdHighest_O3 = ['72', '84.9999']
lowest_O3 = ['62', '71.9999']
# end of O3

# NO2
greaterOrEqual_NO2 = ['90.00001']
secondHighest_NO2 = ['75.00001', '90']
thirdHighest_NO2 = ['60.00001', '75']
lowest_NO2 = ['45.00001', '60']
# end of NO2

# Color codes, it is in HEX
PurpleFill = PatternFill(bgColor="9700d6")
RedFill = PatternFill(bgColor="EE1111")
YellowFill = PatternFill(bgColor="f7ff59")
GreenFill = PatternFill(bgColor="00CE15")

# YYYY/MM/DD
date_style = NamedStyle(name='datetime', number_format='YYYY/MM/DD')
# DO NOT TOUCH THE REST
#####################################################
# gets path of the file
path = os.getcwd()

# auto creates repo
directories = ["O3", "PM25", "NO2"]
if os.path.exists(path + "/monthfiles"):
    shutil.rmtree(path + "/monthfiles")

basedirectories = ["excel_output", "monthfiles", "output"]
for u in basedirectories:
    if os.path.exists(path + "/" + u) is False:
        os.mkdir(path + "/" + u)

for i in directories:
    if not os.path.exists(path + "/monthfiles/"):
        os.mkdir(path + "/" + "monthfiles")
    if not os.path.exists(path + "/monthfiles/" + i):
        os.mkdir(path + "/monthfiles/" + i)

# makes a dictionnary to convert the NAPS and PB
NAPS_PB_File = open("NAPS_PB.csv", "r")
reader = csv.reader(NAPS_PB_File)
converter = list(reader)
EC_Code = []
NAPS_ID = []

for x in range(len(converter)):
    line = converter[x]
    EC = line[0]
    NAPS = line[1]
    EC_Code.append(EC)
    NAPS_ID.append(NAPS)

NAPS_PB_dict = dict(zip(NAPS_ID, EC_Code))
PB_NAPS_dict = dict(zip(EC_Code, NAPS_ID))

# unix path : /fs/home/fs1/ords/oth/airq_central/frc002/Data/CAS/Observations/Station/

# CMVQs
# 50126
print("Enter the start date in YYYY/MM/DD followed by enter")
start = input()
print("Enter the end date in YYYY/MM/DD followed by enter ")
end = input()

startDate = datetime.datetime.strptime(start, "%Y/%m/%d")
endDate = datetime.datetime.strptime(end, "%Y/%m/%d")

# startDate = datetime.datetime(2019, 8, 1)
# endDate = datetime.datetime(2019, 8, 15)

if endDate.date() > datetime.date.today():
    raise Exception('\033[91m' + "Entered End Date is Greater than today" + '\033[0m')

delta = endDate - startDate

filelstCSV = ["O3_" + startDate.strftime("%Y") + "_" + startDate.strftime("%m") + ".csv",
              "NO2_" + startDate.strftime("%Y") + "_" + startDate.strftime("%m") + ".csv",
              "PM25_" + startDate.strftime("%Y") + "_" + startDate.strftime("%m") + ".csv"]

filelstExcel = ["O3_" + startDate.strftime("%Y") + "_" + startDate.strftime("%m") + ".xlsx",
                "NO2_" + startDate.strftime("%Y") + "_" + startDate.strftime("%m") + ".xlsx",
                "PM25_" + startDate.strftime("%Y") + "_" + startDate.strftime("%m") + ".xlsx"]
listofDate = []

for i in range(delta.days + 1):
    listofDate.append((startDate + datetime.timedelta(days=i)).strftime("%Y%m%d"))


# function to find the duplicate and returns the list with their name
def list_duplicates(lst):
    odict = OrderedDict()
    returningOdict = OrderedDict()
    for items in lst:
        odict[items] = None

    for idx, item in enumerate(odict):
        indices = [i for i, x in enumerate(lst) if x == item]
        returningOdict[item] = indices
    return returningOdict


# PM2.5 lists
PM25File = open("StationPM25.csv", "r")
reader = list(csv.reader(PM25File))
PM25_StationIDlst = []
PM25_StationRegionlst = []

for x in range(len(reader)):
    line = reader[x]
    pmID = line[0]
    pmRegion = line[1]
    PM25_StationIDlst.append(pmID)
    PM25_StationRegionlst.append(pmRegion)

PM25_sortedDuplicated = list_duplicates(PM25_StationRegionlst)

PM25indexlist = list(PM25_sortedDuplicated.keys())
PM25_lstDuplicate = list(PM25_sortedDuplicated.values())
###

# O3
O3File = open("StationO3.csv", "r")
Reader_O3 = list(csv.reader(O3File))
O3_StationIDlist = []
O3_StationRegionlst = []

for q in range(len(Reader_O3)):
    l = Reader_O3[q]
    o3ID = l[0]
    o3Region = l[1]
    O3_StationIDlist.append(o3ID)
    O3_StationRegionlst.append(o3Region)

O3_sortedDuplicate = list_duplicates(O3_StationRegionlst)

O3Indexlist = list(O3_sortedDuplicate.keys())
O3_listDuplicate = list(O3_sortedDuplicate.values())
####

# NO2
NO2File = open("StationNO2.csv", "r")
reader_NO2 = list(csv.reader(NO2File))
NO2_stationIDlist = []
NO2_stationRegionlst = []

for a in range(len(reader_NO2)):
    L = reader_NO2[a]
    noID = L[0]
    noRegion = L[1]
    NO2_stationIDlist.append(noID)
    NO2_stationRegionlst.append(noRegion)

NO2_sortedDuplicate = list_duplicates(NO2_stationRegionlst)

NO2Indexlist = list(NO2_sortedDuplicate.keys())
NO2_listDuplicate = list(NO2_sortedDuplicate.values())

######
stationlst = []
datelst = []
hourlst = []
O3lst = []
NO2lst = []
PM25lst = []


# gathers all the files in a montly base
def getPerMonth(lst, polluant):
    for station in lst:
        stationID = PB_NAPS_dict[station]
        if os.path.exists(
                "/fs/home/fs1/ords/oth/airq_central/frc002/Data/CAS/Observations/Station/" + stationID) is True:
            for files in sorted(
                    os.listdir("/fs/home/fs1/ords/oth/airq_central/frc002/Data/CAS/Observations/Station/" + stationID)):
                for days in listofDate:
                    if files.endswith(stationID + "_" + days + ".csv"):
                        f = open(
                            "/fs/home/fs1/ords/oth/airq_central/frc002/Data/CAS/Observations/Station/" + stationID + "/" + files,
                            "r")
                        csvFile = list(csv.reader(f))
                        for z in range(len(csvFile)):
                            row = csvFile[z]
                            if row != ['station', 'Date', 'UTC', 'AQHI', 'O3', 'NO2', 'PM2.5', 'PM10', 'SO2', 'H2S',
                                       'CO',
                                       'NO', 'TRS']:
                                station = row[0]
                                stationlst.append(station)
                                date = row[1]
                                datelst.append(date)
                                hour = row[2]
                                hourlst.append(hour)
                                NO2 = row[5]
                                NO2lst.append(NO2)
                                O3 = row[4]
                                O3lst.append(O3)
                                PM25 = row[6]
                                PM25lst.append(PM25)

        if polluant is 1:
            df2 = pd.DataFrame(data={"Date(DD/MM/YYYY)": datelst, "Concentration": PM25lst, "Hour(UTC)": hourlst})
            df2.to_csv("monthfiles/PM25/" + startDate.strftime("%Y%m") + "_" + stationID + "PM25.csv", sep=",",
                       index=False)

        if polluant is 2:
            df1 = pd.DataFrame(data={"Date(DD/MM/YYYY)": datelst, "Concentration": NO2lst, "Hour(UTC)": hourlst})
            df1.to_csv("monthfiles/NO2/" + startDate.strftime("%Y%m") + "_" + stationID + "NO2.csv", sep=",",
                       index=False)

        if polluant is 3:
            df = pd.DataFrame(data={"Date(DD/MM/YYYY)": datelst, "Concentration": O3lst, "Hour(UTC)": hourlst})
            df.to_csv("monthfiles/O3/" + startDate.strftime("%Y%m") + "_" + stationID + "O3.csv", sep=",", index=False)

        monthtemplate = pd.DataFrame(data={"Date(DD/MM/YYYY)": datelst, "Hour(UTC)": hourlst})
        monthtemplate.to_csv("monthTemplate.csv", sep=",", index=False)
        stationlst.clear()
        datelst.clear()
        hourlst.clear()
        NO2lst.clear()
        O3lst.clear()
        PM25lst.clear()


# generates montly report based on station list (to keep order)
def generateMonthReport(stat, polluant):
    template = open("monthTemplate.csv", "r")
    csvr = list(csv.reader(template))
    hlst = []
    daylst = []
    for q in range(len(csvr)):
        r = csvr[q]
        d = r[0]
        h = r[1]
        daylst.append(d + " " + h)
        hlst.append(h)

    cclst = []

    finalFile = pd.DataFrame(data={"Date(DD/MM/YYYY)": daylst[1:], "Hours(UTC)": hlst[1:]})
    for station in stat:
        stationID = PB_NAPS_dict[station]
        for file in os.listdir("monthfiles/" + polluant):
            if file.endswith(stationID + polluant + ".csv"):
                F = open("monthfiles/" + polluant + "/" + file, "r")
                CSVFiles = list(csv.reader(F))
                for t in range(len(CSVFiles)):
                    rows = CSVFiles[t]
                    concentration = rows[0]
                    cclst.append(concentration)
                    # print(rows)
        # print(len(cclst[1:]), len(daylst[1:]))
        if len(cclst[1:]) != len(daylst[1:]):
            continue
        finalFile[station] = cclst[1:]
        finalFile.to_csv(
            "output/" + polluant + "_" + startDate.strftime("%Y") + "_" + startDate.strftime("%m") + ".csv", sep=",",
            index=False)
        cclst.clear()


def convertToExcel(filelst):
    for files in filelst:
        # for files in os.listdir("output"):
        csvIn = pd.read_csv("output/" + files, delimiter=",").replace(to_replace="-9.99", value="")
        # converts to datetime object
        csvIn['Date(DD/MM/YYYY)'] = pd.to_datetime(csvIn['Date(DD/MM/YYYY)'], infer_datetime_format=True)
        excelout = pd.ExcelWriter("excel_output/" + files.split(".")[0] + ".xlsx", engine='xlsxwriter')
        csvIn.to_excel(excelout, sheet_name="Original Data", index=False)
        wbook = excelout.book
        ws = wbook.add_worksheet('3h Average')
        ws1 = wbook.add_worksheet('Regional Hour Max')
        ws2 = wbook.add_worksheet('Everything Together')
        excelout.save()


def addcolor(sheet, totalcolumns, numberRow, firstbound, secondbound, thirdbound, fourthbound):
    sheet.conditional_formatting.add('C2:' + get_column_letter(totalcolumns) + str(numberRow),
                                     CellIsRule(operator='greaterThanOrEqual', formula=firstbound,
                                                stopIfTrue=True,
                                                fill=PurpleFill))

    sheet.conditional_formatting.add('C2:' + get_column_letter(totalcolumns) + str(numberRow),
                                     CellIsRule(operator='between', formula=secondbound,
                                                stopIfTrue=True,
                                                fill=RedFill))

    sheet.conditional_formatting.add('C2:' + get_column_letter(totalcolumns) + str(numberRow),
                                     CellIsRule(operator='between', formula=thirdbound,
                                                stopIfTrue=True,
                                                fill=YellowFill))

    sheet.conditional_formatting.add('C2:' + get_column_letter(totalcolumns) + str(numberRow),
                                     CellIsRule(operator='between', formula=fourthbound, stopIfTrue=True,
                                                fill=GreenFill))


# formula to calculate 3h avg


def Avg3handMax(excelFiles, startStr, indexList, firstbound, secondbound, thirdbound, fourthbound, listduplicate):
    # for excelFiles in os.listdir("excel_output"):
    wb = openpyxl.load_workbook("excel_output/" + excelFiles)
    rawdata = wb['Original Data']
    avg_3h = wb['3h Average']
    regionMax = wb['Regional Hour Max']
    everything = wb['Everything Together']
    numberRow = rawdata.max_row
    numberColumn = rawdata.max_column
    regionalMaxcolum = len(indexList) + 2

    rawdata_delta = rawdata.max_column + 1
    regional_delta = regionalMaxcolum + 2

    # creates graph for region max
    # hourly reg max: get_column_letter(len(indexList) + 4) + str(r)

    concentration_data = Reference(regionMax, min_col=len(indexList) + 4, min_row=1, max_row=numberRow)
    dates_data = Reference(regionMax, min_col=1, min_row=2, max_row=numberRow)

    qc24regionmax_chart = ScatterChart()
    qc24regionmax_chart.title = "Quebec Monthly Max Graph"
    qc24regionmax_chart.style = 2
    qc24regionmax_chart.y_axis.title = "Concentration"
    qc24regionmax_chart.x_axis.number_format = 'dd'
    qc24regionmax_chart.x_axis.majorTimeUnit = "days"
    qc24regionmax_chart.x_axis.title = "Date"

    qc24hSeries = Series(concentration_data, dates_data, title_from_data=True)
    qc24regionmax_chart.series.append(qc24hSeries)
    regionMax.add_chart(qc24regionmax_chart, get_column_letter(len(indexList) + 6) + "1")

    for sheets in wb.worksheets:
        sheets.freeze_panes = "C2"
        # copies row/column from old set
        if sheets != wb['Original Data']:
            for r in range(1, numberRow + 1):
                for c in range(1, 3):
                    # write the stations
                    if r is 1:
                        for allC in range(1, numberColumn + 1):

                            # creates the 3h avg station row header
                            if sheets == avg_3h:
                                sheets[get_column_letter(allC) + str(r)] \
                                    = '=(\'Original Data\'!' + get_column_letter(allC) + str(r) + ')'

                        # write region max row header
                        if excelFiles.startswith(startStr):
                            for i, p in enumerate(indexList):
                                if sheets == regionMax:
                                    sheets[get_column_letter(i + 3) + str(r)] = p

                    # write time and date for both 3h avg and region max sheet
                    if r is 2 or r is 3:
                        for allC in range(3, numberColumn + 1):
                            sheets[get_column_letter(allC) + str(r)] = ''
                    sheets[get_column_letter(c) + str(r)] = '=(\'Original Data\'!' + get_column_letter(c) + str(
                        r) + ')'

        # applies conditional formatting rule
        if excelFiles.startswith(startStr):
            if sheets == everything:
                if startStr == "PM25":
                    addcolor(everything, 2 * rawdata_delta + regional_delta, numberRow, firstbound, secondbound,
                             thirdbound, fourthbound)
                if startStr == ("NO2" or "O3"):
                    addcolor(everything, rawdata_delta + regional_delta, numberRow, firstbound, secondbound, thirdbound,
                             fourthbound)

            if sheets != everything:
                addcolor(sheets, numberColumn, numberRow, firstbound, secondbound, thirdbound,
                         fourthbound)

    # rounds everything to integer
    for i in range(0, numberRow - 3):
        for o in range(3, numberColumn + 1):
            avg_3h[get_column_letter(o) + str(i + 4)] = \
                '=IF((COUNTA(\'Original Data\'!' + get_column_letter(o) + str(i + 2) + ':' + get_column_letter(o) \
                + str(i + 4) + '))>1,ROUND(AVERAGE(\'Original Data\'!' + \
                get_column_letter(o) + str(i + 2) + ':' + get_column_letter(o) + str(i + 4) + '),0),-999)'

    # calculate region max, this one has to be based on 3h avg
    if startStr == "PM25":
        regionMax[get_column_letter(len(indexList) + 4) + '1'] = "Hour QC Max"
        for r in range(4, numberRow + 1):
            for c, data in zip(range(3, len(indexList) + 3), listduplicate):
                sData = data[0] + 3
                endData = data[-1] + 3
                regionMax[get_column_letter(c) + str(r)] = '=MAX(\'3h Average\'!' \
                                                           + get_column_letter(sData) \
                                                           + str(r) + ':' + \
                                                           get_column_letter(endData) \
                                                           + str(r) + ')'
            # for the hourly max, it has to exclude the temis region
            for name in indexList:
                if name == "Temis.":
                    temisLetter = get_column_letter(indexList.index("Temis.") + 3)
                    regionMax[get_column_letter(len(indexList) + 4) + str(r)] \
                        = '=IF(MAX(' + get_column_letter(3) + str(r) + ':' \
                          + get_column_letter(regionalMaxcolum) + str(r) + ')=' \
                          + temisLetter + str(r) + ',LARGE(' + get_column_letter(3) + str(r) + ':' \
                          + get_column_letter(regionalMaxcolum) + str(r) + ',2),MAX(' + get_column_letter(3) + str(
                        r) + ':' \
                          + get_column_letter(regionalMaxcolum) + str(r) + '))'

        # This part of script rewrite all data into one sheet
        for r in range(1, rawdata.max_row + 1):
            for c in range(1, rawdata.max_column + 1):
                # ROUND(\'Original Data\'!' + get_column_letter(c) + str(r) + ',0))
                everything[get_column_letter(c) + str(r)] = '=IF((\'Original Data\'!' + get_column_letter(c) + str(
                    r) + ')="",-999,(\'Original Data\'!' + get_column_letter(c) + str(r) + '))'

        # 3h avgs copy
        delta = rawdata.max_column + 1
        for avg_C in range(3, avg_3h.max_column + 1):
            for avg_R in range(1, avg_3h.max_row + 1):
                everything[get_column_letter(avg_C + delta) + str(avg_R)] = '=IF((\'3h Average\'!' + get_column_letter(
                    avg_C) + str(
                    avg_R) + ')="",-999,(\'3h Average\'!' + get_column_letter(avg_C) + str(avg_R) + '))'

        # reg hour max copy
        avg3h_delta = avg_3h.max_column
        for reg_C in range(3, len(indexList) + 5):
            for reg_R in range(1, regionMax.max_row + 1):
                everything[get_column_letter(reg_C + delta + avg3h_delta) + str(
                    reg_R)] = '=IF((\'Regional Hour Max\'!' + get_column_letter(reg_C) + str(
                    reg_R) + ')="",-999,(\'Regional Hour Max\'!' + get_column_letter(reg_C) + str(reg_R) + '))'



    else:
        # NO2 and O3 are based on direct observation
        # a bit of a hack but 3h avg are not needed for them so I just remove the sheet
        wb.remove_sheet(avg_3h)
        regionMax[get_column_letter(len(indexList) + 4) + '1'] = "Hour QC Max"
        for r in range(2, numberRow + 1):
            for c, data in zip(range(3, len(indexList) + 3), listduplicate):
                sData = data[0] + 3
                endData = data[-1] + 3
                regionMax[get_column_letter(c) + str(r)] = '=MAX(\'Original Data\'!' \
                                                           + get_column_letter(sData) \
                                                           + str(r) + ':' + \
                                                           get_column_letter(endData) \
                                                           + str(r) + ')'
            # for hourly max
            regionMax[get_column_letter(len(indexList) + 4) + str(r)] = '=MAX(' + get_column_letter(3) + str(
                r) + ':' \
                                                                        + get_column_letter(regionalMaxcolum) + str(
                r) + ')'

        # This part of script rewrite all data into one sheet
        for r in range(1, rawdata.max_row + 1):
            for c in range(1, rawdata.max_column + 1):
                everything[get_column_letter(c) + str(r)] = '=IF((\'Original Data\'!' + get_column_letter(c) + str(
                    r) + ')="",-999,(\'Original Data\'!' + get_column_letter(c) + str(r) + '))'

        # reg hour max copy
        delta = rawdata.max_column + 1
        for reg_C in range(3, len(indexList) + 5):
            for reg_R in range(1, regionMax.max_row + 1):
                everything[get_column_letter(reg_C + delta) + str(
                    reg_R)] = '=IF((\'Regional Hour Max\'!' + get_column_letter(reg_C) + str(
                    reg_R) + ')="",-999,(\'Regional Hour Max\'!' + get_column_letter(reg_C) + str(reg_R) + '))'

    # write daily max for all 3 file
    regionMax[get_column_letter(len(indexList) + 5) + '1'] = "Daily Max"
    for day in range(len(listofDate)):
        for r in range(4, numberRow + 1):
            regionMax[get_column_letter(len(indexList) + 5) + str(2 + day * 24)] = '=MAX(' + get_column_letter(
                len(indexList) + 4) + str(2 + day * 24) + ':' + get_column_letter(len(indexList) + 4) + str(
                25 + day * 24) + ')'
    # does it for everything sheet
    dailyMaxLetter = get_column_letter(everything.max_column + 1)
    qcHourMaxLetter = get_column_letter(everything.max_column)
    everything[dailyMaxLetter + '1'] = "Daily Max"
    for day in range(len(listofDate)):
        for r in range(4, numberRow + 1):
            everything[dailyMaxLetter + str(2 + day * 24)] = '=MAX(' + qcHourMaxLetter + str(
                2 + day * 24) + ':' + qcHourMaxLetter + str(
                25 + day * 24) + ')'
    wb.save("excel_output/" + excelFiles)


# these are the functions calls. It reduce the code bulkiness
print("Starting Request...")
getPerMonth(PM25_StationIDlst, 1)
print("25% done")
getPerMonth(NO2_stationIDlist, 2)
getPerMonth(O3_StationIDlist, 3)
generateMonthReport(NO2_stationIDlist, "NO2")
generateMonthReport(O3_StationIDlist, "O3")
print("50% done")
generateMonthReport(PM25_StationIDlst, "PM25")
convertToExcel(filelstCSV)
Avg3handMax(filelstExcel[0], "O3", O3Indexlist, greaterOrEqual_O3, secondHighest_O3, thirdHighest_O3, lowest_O3,
            O3_listDuplicate)
print("75% done")
Avg3handMax(filelstExcel[1], "NO2", NO2Indexlist, greaterOrEqual_NO2, secondHighest_NO2, thirdHighest_NO2, lowest_NO2,
            NO2_listDuplicate)
Avg3handMax(filelstExcel[2], "PM25", PM25indexlist, greaterOrEqual_PM25, secondHighest_PM25, thirdHighest_PM25,
            lowest_PM25, PM25_lstDuplicate)

print("\033[32m" + "100% :)")
print("Job done, see --> " + path + "/excel_output" + '\033[0m')
