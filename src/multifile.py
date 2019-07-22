from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill

wb = Workbook()
ws = wb.active
RedFill = PatternFill(bgColor="4A009E")
# Create fill
# start_color='EE1111',end_color="EE1111",fill_type='solid'

# Add a two-color scale

# Add a conditional formatting based on a cell comparison
# addCellIs(range_string, operator, formula, stopIfTrue, wb, font, border, fill)
# Format if cell is less than 'formula'

# Format if cell is between 'formula'
ws.conditional_formatting.add('D2:D10',
                              CellIsRule(operator='between', formula=['1', '5'], stopIfTrue=True, fill=RedFill))

wb.save("test.xlsx")
