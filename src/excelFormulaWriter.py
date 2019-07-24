import csv
import os
import time
from collections import defaultdict

import openpyxl
import pandas as pd
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill
from openpyxl.utils.cell import get_column_letter

s = time.time()

PM25File = open("StationPM25.csv", "r")
reader = list(csv.reader(PM25File))
PM25_StationIDlst = []
PM25_StationRegionlst = []

for x in range(len(reader)):
    line = reader[x]
    pmID = line[0]
    pmRegion = line[1]
    PM25_StationIDlst.append(pmID)
    PM25_StationRegionlst.append(pmRegion)

tempset = set()
regionDictionary = dict(zip(PM25_StationIDlst, PM25_StationRegionlst))
for w in PM25_StationIDlst:
    a = regionDictionary[w]
    tempset.add(PM25_StationRegionlst.index(a))

PM25indexlist = list(sorted(tempset))


# converts to excel file
for files in os.listdir("output"):
    csvIn = pd.read_csv("output/" + files, delimiter=",")
    excelout = pd.ExcelWriter("excel_output/" + files.split(".")[0] + ".xlsx", engine='xlsxwriter')
    csvIn.to_excel(excelout, sheet_name="Original Data", index=False)
    # csvIn.to_excel(excelout, sheet_name="3h Average", index=False)
    wb = excelout.book
    ws = wb.add_worksheet('3h Average')
    ws1 = wb.add_worksheet('Regional Hour Max')
    excelout.save()


def list_duplicates(seq):
    tally = defaultdict(list)
    for i, item in enumerate(seq):
        tally[item].append(i)
    return ((key, locs) for key, locs in tally.items())


PM25_lstDuplicate = []
for region in list_duplicates(PM25_StationRegionlst):
    PM25_lstDuplicate.append(region[1])

print(len(PM25_lstDuplicate))
for excelFiles in os.listdir("excel_output"):

    PurpleFill = PatternFill(bgColor="56007a")
    RedFill = PatternFill(bgColor="EE1111")
    YellowFill = PatternFill(bgColor="EECE00")
    GreenFill = PatternFill(bgColor="00CE15")

    wb = openpyxl.load_workbook("excel_output/" + excelFiles)
    rawdata = wb['Original Data']
    avg_3h = wb['3h Average']
    regionMax = wb['Regional Hour Max']
    numberRow = rawdata.max_row
    numberColumn = rawdata.max_column

    for sheets in wb.worksheets:
        # copies row/column from old set
        if sheets != wb['Original Data']:
            for r in range(1, numberRow + 1):
                for c in range(1, 3):
                    if r is 1:
                        for allC in range(1, numberColumn + 1):

                            if sheets == avg_3h:
                                sheets[get_column_letter(allC) + str(r)] = '=(\'Original Data\'!' + \
                                                                           get_column_letter(allC) + str(r) + ')'
                        ##
                        if excelFiles.startswith("PM25"):  ##
                            for i, p in enumerate(PM25indexlist):
                                if sheets == regionMax:
                                    # print(get_column_letter(i+3) + str(r), PM25_StationRegionlst[p])
                                    sheets[get_column_letter(i + 3) + str(r)] = PM25_StationRegionlst[p]
                    if r is 2 or r is 3:
                        for allC in range(3, numberColumn + 1):
                            sheets[get_column_letter(allC) + str(r)] = ''
                    sheets[get_column_letter(c) + str(r)] = '=(\'Original Data\'!' + get_column_letter(c) + str(r) + ')'

        # conditional formatting
        if excelFiles.startswith("PM25"):
            sheets.conditional_formatting.add('C2:' + get_column_letter(numberColumn) + str(numberRow),
                                              CellIsRule(operator='greaterThanOrEqual', formula=['50'], stopIfTrue=True,
                                                         fill=PurpleFill))

            sheets.conditional_formatting.add('C2:' + get_column_letter(numberColumn) + str(numberRow),
                                              CellIsRule(operator='between', formula=['35', '49.9999'], stopIfTrue=True,
                                                         fill=RedFill))

            sheets.conditional_formatting.add('C2:' + get_column_letter(numberColumn) + str(numberRow),
                                              CellIsRule(operator='between', formula=['30', '34.9999'], stopIfTrue=True,
                                                         fill=YellowFill))

            sheets.conditional_formatting.add('C2:' + get_column_letter(numberColumn) + str(numberRow),
                                              CellIsRule(operator='between', formula=['25', '29.9999'], stopIfTrue=True,
                                                         fill=GreenFill))

        if excelFiles.startswith("O3"):
            sheets.conditional_formatting.add('C2:' + get_column_letter(numberColumn) + str(numberRow),
                                              CellIsRule(operator='greaterThanOrEqual', formula=['100'],
                                                         stopIfTrue=True,
                                                         fill=PurpleFill))

            sheets.conditional_formatting.add('C2:' + get_column_letter(numberColumn) + str(numberRow),
                                              CellIsRule(operator='between', formula=['85', '99.9999'], stopIfTrue=True,
                                                         fill=RedFill))

            sheets.conditional_formatting.add('C2:' + get_column_letter(numberColumn) + str(numberRow),
                                              CellIsRule(operator='between', formula=['72', '84.9999'], stopIfTrue=True,
                                                         fill=YellowFill))

            sheets.conditional_formatting.add('C2:' + get_column_letter(numberColumn) + str(numberRow),
                                              CellIsRule(operator='between', formula=['62', '71.9999'], stopIfTrue=True,
                                                         fill=GreenFill))

            # todo add for NO2

    # formulae for 3h avg
    for i in range(0, numberRow - 3):
        for o in range(3, numberColumn + 1):
            avg_3h[get_column_letter(o) + str(i + 4)] = '=IF((COUNTA(\'Original Data\'!' + get_column_letter(o) + str(
                i + 2) + ':' + get_column_letter(o) \
                                                        + str(i + 4) + '))>1,ROUND(AVERAGE(ROUND(\'Original Data\'!' + \
                                                        get_column_letter(o) + str(i + 2) + \
                                                        ',0),ROUND(\'Original Data\'!' + get_column_letter(o) + str(
                i + 3) + \
                                                        ',0),ROUND(\'Original Data\'!' + get_column_letter(o) + str(
                i + 4) + ',0)),0),\" \")'

    # regional max
    if excelFiles.startswith("PM25"):
        print(excelFiles)
        for r in range(4, numberRow + 1):
            for c, data in zip(range(3, len(PM25indexlist) + 3), PM25_lstDuplicate):
                sData = data[0] + 3
                endData = data[-1] + 3
                print(get_column_letter(c), r, sData, endData)
                regionMax[get_column_letter(c) + str(r)] = '=MAX(\'3h Average\'!' \
                                                           + get_column_letter(sData) \
                                                           + str(r) + ':' + \
                                                           get_column_letter(endData) \
                                                           + str(r) + ')'

                # print(get_column_letter(c),c,get_column_letter(r),r)

        #           s= region[1][0]+3
        #           e= region[1][-1]+3
        #           print(s,e)
        #           print(get_column_letter(s),get_column_letter(e),o,r)
        #           print("--")
        #           # except:
        #           #     PM25indexlist[-1] = PM25indexlist[i+1]
        #           #     print(PM25indexlist[i],PM25indexlist[i + 1])
        #           #     continue

        # #
    wb.save("excel_output/" + excelFiles)

e = time.time()
print(e - s)
# "=IF((COUNTA(raw!C2:C4))>1,(AVERAGE(ROUND(raw!C2,0),ROUND(raw!C3,0),ROUND(raw!C4,0))),\" \")"
# https://automatetheboringstuff.com/chapter12/
